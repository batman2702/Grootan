from selenium import webdriver
import time
import xlsxwriter
import openpyxl as xl
import os
from PIL import Image
from selenium.common.exceptions import NoSuchElementException

class Groot:
    path = "C:\\Users\\Administrator\\Downloads\\geckodriver.exe"
    row=1
    test=1
    def __init__(self, url):
        self.url = url
        self.driver = webdriver.Firefox(executable_path=Groot.path)
        workbook = xlsxwriter.Workbook("Reports.xlsx")
        worksheet = workbook.add_worksheet('TSR')
        worksheet2 = workbook.add_worksheet('JuniorEngineers')
        worksheet.set_column(0, 0, 30)
        worksheet.set_column(2, 2, 50)
        worksheet.set_column(4, 4, 10)
        worksheet.set_column(6, 6, 20)
        worksheet.write('A1','TestCase')
        worksheet.write('C1','Description')
        worksheet.write('E1','Status')
        worksheet.write('G1','Screenshot')
        worksheet2.write(0,0,30)
        worksheet2.write('A1','Name')
        workbook.close()

    def functional_testing(self):
        os.makedirs("Folder" + str(Groot.test))
        cwd = os.getcwd()
        path = os.path.join(cwd, 'Folder' + str(Groot.test))
        Groot.test+=1
        self.driver.get(self.url)
        self.driver.maximize_window()
        self.driver.save_screenshot("root.png")
        self.excelupdation("Testcase1","Get request to grootan website","Pass","root.png")
        time.sleep(15)
        try:
            ele=self.driver.find_element_by_xpath("//footer//div//div//div[2]//button[@type='button']")
            ele.click()
        except NoSuchElementException:
            pass
        time.sleep(10)
        self.driver.save_screenshot(path+"\\Home.png")
        self.driver.find_element_by_xpath("//nav[@id='main-nav']//ul//a[@href='/#built-tech']").click()
        time.sleep(5)
        self.driver.save_screenshot(path + "\\Services.png")
        self.driver.find_element_by_xpath("//nav[@id='main-nav']//div//ul//li[2]//a[@href='/opensource']").click()
        time.sleep(5)
        self.driver.save_screenshot(path+"\\Opensource.png")
        self.driver.find_element_by_xpath("//nav[@id='main-nav']//div//ul//li[2]//a[4]").click()
        time.sleep(5)
        self.driver.save_screenshot(path+"\\Blog.png")
        self.driver.find_element_by_xpath("//div[@id='gatsby-focus-wrapper']//nav[@role='navigation']//ul//a[@href='https://www.grootan.com/team']").click()
        time.sleep(5)
        self.driver.save_screenshot(path+"\\Team.png")
        self.driver.find_element_by_xpath("//nav[@id='main-nav']//div//ul//li[2]//a[@href='/careers']").click()
        time.sleep(5)
        self.driver.save_screenshot(path+"\\Careers.png")
        self.driver.find_element_by_xpath("//nav[@id='main-nav']//div//ul//li[2]//a[@href='/contactus']").click()
        time.sleep(5)
        self.driver.save_screenshot(path+"\\Contactus.png")


    def excelupdation(self,name,description,status,image):
        xfile = xl.load_workbook("Reports.xlsx")
        sheet = xfile['TSR']
        sheet['A'+str(Groot.row+1)]=name
        sheet['C'+str(Groot.row+1)] = description
        sheet['E' + str(Groot.row + 1)] = status

        if image not in(0,1):
            img = xl.drawing.image.Image(image)
            img.height = 250
            img.width = 300
            img.anchor = "G"+str(Groot.row+1)
            sheet.add_image(img)
        else:
            if image==0:
                sheet['G' + str(Groot.row + 1)] = "Same images"
                #sheet.range('E' + str(Groot.row + 1)).value = "Same images"
            else:
                sheet['G' + str(Groot.row + 1)] = "Different Images images"
                #sheet.range('E' + str(Groot.row + 1)).value = "Different Images images"

        Groot.row+=13
        xfile.save("Reports.xlsx")

    def screenshot_comparison(self):
        ls = ["Home.png","Opensource.png","Blog.png","Team.png","Careers.png","Contactus.png"]
        cwd = os.getcwd()
        path1 = os.path.join(cwd, "Folder1")
        path2 = os.path.join(cwd, "Folder2")
        for pic in ls:
            image_one = Image.open(path1 + "\\" + pic)
            for pic1 in ls:
                image_two = Image.open(path2 + "\\"+pic1)
                if list(image_one.getdata()) != list(image_two.getdata()):
                    self.excelupdation("Screenshot Comparison","Comparing Folder1,Folder2 screenshots","Fails",1)
                else:
                    self.excelupdation("Screenshot Comparison", "Comparing Folder1,Folder2 screenshots", "Success", 0)

    def imagecomparison(self):
        self.driver.get("https://www.grootan.com/team")
        self.driver.maximize_window()
        ls = self.driver.find_elements_by_css_selector("[class='col-lg-3 col-sm-6']")
        url, name, des = [], [], []
        temp1=1
        for i in range(len(ls) - 1):
            tem = ls[i].find_element_by_tag_name('img')
            tem = tem.get_attribute('src')
            url.append(tem)
            temp = tuple(ls[i].text.split('\n'))
            if len(temp)>1 and temp[1] == 'Junior Engineer':
                xfile = xl.load_workbook("Reports.xlsx")
                sheet=xfile['JuniorEngineers']
                sheet['A' + str(temp1)] = temp[0]
                xfile.save("Reports.xlsx")
                temp1+=1
            name.append(temp[0])
            des.append(temp[1])
        if url[0]!=url[1]:
            self.excelupdation("Image comparison","Comparing CTO and HR images","Fail",1)
        else:
            self.excelupdation("Image comparison", "Comparing CTO and HR images", "Success", 0)




    def driver_close(self):
        self.driver.close()



# driver=webdriver.Chrome(executable_path="C:\\Users\\Administrator\\Downloads\\chromedriver.exe")
#
# driver.get("https://www.grootan.com/")
# time.sleep(5)
# driver.save_screenshot("home.png")
# driver.find_element_by_xpath("//div[@id='root']//div//nav//div//ul//li[2]//a").click()
#
# driver.find_element_by_xpath("//div[@id='root']//div//nav//div//ul//li[2]//a[2]").click()
# time.sleep(5)
# driver.save_screenshot("services.png")
#
# driver.find_element_by_xpath("//div[@id='root']//div//nav//div//ul//li[2]//a[2]").click()
# time.sleep(5)
# driver.save_screenshot("services.png")



#driver.close()

obj = Groot("https://www.grootan.com/")
obj.functional_testing()
obj.functional_testing()
obj.screenshot_comparison()
obj.imagecomparison()
obj.driver_close()

